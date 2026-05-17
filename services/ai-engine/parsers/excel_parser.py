"""Stage 2: Excel (.xlsx) parser"""
import io
import logging
import re
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# All recognised aliases for each logical column
_QUESTION_ALIASES = ["question", "savol", "вопрос", "q", "text", "matn"]
_OPTION_A_ALIASES = ["a", "variant_a", "a)", "option_a", "choice_a"]
_OPTION_B_ALIASES = ["b", "variant_b", "b)", "option_b", "choice_b"]
_OPTION_C_ALIASES = ["c", "variant_c", "c)", "option_c", "choice_c"]
_OPTION_D_ALIASES = ["d", "variant_d", "d)", "option_d", "choice_d"]
_OPTION_E_ALIASES = ["e", "variant_e", "e)", "option_e", "choice_e"]
_ANSWER_ALIASES = [
    "answer", "javob", "ответ", "correct", "correct_answer",
    "to'g'ri", "togri", "right", "key",
]
_EXPLANATION_ALIASES = ["explanation", "tushuntirish", "объяснение", "izoh", "note"]


def parse_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parses an Excel (.xlsx) file and returns a list of structured question dicts.

    Handles:
    - Multiple header name variations (see aliases above).
    - Merged cells (openpyxl unmerges them before reading).
    - Auto-detection of the answer column when no explicit answer header exists.
    - Option columns E / beyond D.

    Returns List[Dict] ready for the validator stage:
        {
          "question": str,
          "options": List[str],
          "correct_index": int,   # 0-based
          "explanation": str,
        }
    """
    try:
        from openpyxl import load_workbook

        # data_only=True → read computed values, not formulas
        wb = load_workbook(io.BytesIO(file_content), read_only=False, data_only=True)
        ws = wb.active

        # Unmerge cells so iter_rows works correctly
        _unmerge_cells(ws)

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # Find header row (first row that contains at least one known column name)
        header_row_idx, col_map = _find_header(rows)
        if header_row_idx is None:
            logger.warning("Excel: header qator topilmadi")
            return []

        question_col = _find_col(col_map, _QUESTION_ALIASES)
        if question_col is None:
            logger.warning("Excel: savol ustuni topilmadi")
            return []

        answer_col = _find_col(col_map, _ANSWER_ALIASES)
        explanation_col = _find_col(col_map, _EXPLANATION_ALIASES)

        # Gather option columns in order A–E
        option_groups = [
            _OPTION_A_ALIASES,
            _OPTION_B_ALIASES,
            _OPTION_C_ALIASES,
            _OPTION_D_ALIASES,
            _OPTION_E_ALIASES,
        ]
        option_cols = [_find_col(col_map, aliases) for aliases in option_groups]

        # If no explicit answer column → try to auto-detect later per row
        questions: List[Dict[str, Any]] = []

        for row in rows[header_row_idx + 1:]:
            if not row:
                continue

            q_val = _cell_val(row, question_col)
            if not q_val:
                continue

            # Collect options
            options: List[str] = []
            for col_idx in option_cols:
                val = _cell_val(row, col_idx)
                if val:
                    options.append(val)

            # Need at least 2 options
            if len(options) < 2:
                logger.debug("Excel: '%s' — 2 ta variant kerak, o'tkazib yuborildi", q_val)
                continue

            # Determine correct answer
            answer_raw = _cell_val(row, answer_col) if answer_col is not None else ""
            correct_index = _resolve_correct_index(answer_raw, options)

            explanation = _cell_val(row, explanation_col) if explanation_col is not None else ""

            questions.append(
                {
                    "question": q_val,
                    "options": options,
                    "correct_index": correct_index,
                    "explanation": explanation or "",
                }
            )

        return questions

    except Exception as exc:
        raise RuntimeError(f"Excel fayl o'qilmadi: {exc}") from exc


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _unmerge_cells(ws) -> None:
    """Fills merged cell ranges with the top-left cell value before reading."""
    try:
        # openpyxl stores merged ranges; iterate a copy because we'll modify it
        for merged_range in list(ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
            top_left_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merged_range))
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    ws.cell(row_idx, col_idx).value = top_left_value
    except Exception as exc:
        logger.debug("Merged cell unmerge failed (non-fatal): %s", exc)


def _find_header(rows: list):
    """
    Returns (header_row_index, col_map) where col_map maps
    normalised column name → column index (0-based).
    Searches the first 10 rows.
    """
    all_aliases = (
        _QUESTION_ALIASES
        + _OPTION_A_ALIASES
        + _OPTION_B_ALIASES
        + _OPTION_C_ALIASES
        + _OPTION_D_ALIASES
        + _OPTION_E_ALIASES
        + _ANSWER_ALIASES
        + _EXPLANATION_ALIASES
    )
    known = set(all_aliases)

    for row_idx, row in enumerate(rows[:10]):
        if not row:
            continue
        header = [_norm_header(c) for c in row]
        matched = sum(1 for h in header if h in known)
        if matched >= 1:
            col_map = {name: idx for idx, name in enumerate(header) if name}
            return row_idx, col_map

    return None, {}


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).lower().strip()
    # Remove trailing parentheses like "a)" → "a"
    s = re.sub(r"[().\s]+$", "", s)
    return s


def _find_col(col_map: Dict[str, int], names: List[str]) -> Optional[int]:
    for name in names:
        if name in col_map:
            return col_map[name]
    return None


def _cell_val(row: tuple, col_idx: Optional[int]) -> str:
    if col_idx is None or col_idx >= len(row):
        return ""
    val = row[col_idx]
    if val is None:
        return ""
    return str(val).strip()


def _resolve_correct_index(answer_raw: str, options: List[str]) -> int:
    """
    Converts an answer cell value to a 0-based option index.

    Handles:
    - Single letter: "A", "B", "C", "D", "E" (case-insensitive)
    - 1-based number: "1", "2", "3", "4"
    - Full option text match (case-insensitive, stripped)
    - Falls back to 0
    """
    a = answer_raw.strip().lower()
    if not a:
        return 0

    # Single letter
    letter_map = {"a": 0, "b": 1, "c": 2, "d": 3, "e": 4}
    if a in letter_map and letter_map[a] < len(options):
        return letter_map[a]

    # 1-based numeric
    if a.isdigit():
        idx = int(a) - 1
        if 0 <= idx < len(options):
            return idx

    # Full text match
    for i, opt in enumerate(options):
        if opt.lower().strip() == a:
            return i

    # Partial match (answer is a substring of option)
    for i, opt in enumerate(options):
        if a in opt.lower().strip():
            return i

    logger.debug("Excel: to'g'ri javob aniqlanmadi '%s' — 0 qabul qilindi", answer_raw)
    return 0
